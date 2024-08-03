import pandas as pd
import dash
import dash_core_components as dcc
import dash_html_components as html
import plotly.express as px
from sklearn.linear_model import LinearRegression
import numpy as np
from openpyxl import load_workbook

# Initialize the Dash app
app = dash.Dash(__name__)

# File path to the Excel file
file_path = 'C:/Users/27643/Downloads/Chart in Microsoft PowerPoint.xlsx'

# Load the workbook and select sheets
workbook = load_workbook(file_path, data_only=True)
sheet_names = workbook.sheetnames

# Extract data from each sheet
titles = {}
x_labels = {}
y_labels = {}
dfs = {}

for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    titles[sheet_name] = sheet['A1'].value
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=0)
    x_labels[sheet_name] = df.iloc[1, 0]  # X-axis label from A2
    y_labels[sheet_name] = df.iloc[1, 1]  # Y-axis label from B2
    dfs[sheet_name] = df.iloc[2:, [0, 1]].rename(columns={0: x_labels[sheet_name], 1: y_labels[sheet_name]})

# Create bar charts
def create_bar_chart(df, title, x_label, y_label, color):
    fig = px.bar(
        df,
        x=df.columns[0],
        y=df.columns[1],
        title=title,
        labels={df.columns[0]: x_label, df.columns[1]: y_label},
        color_discrete_sequence=[color]
    )
    fig.update_layout(
        autosize=False,
        width=600,
        height=600,
        bargap=0.2
    )
    fig.update_traces(text=df[df.columns[1]], textposition='outside')  # Add values on bars
    return fig

# Create line charts
def create_line_chart(df, title, x_label, y_label, color):
    fig = px.line(
        df,
        x=df.columns[0],
        y=df.columns[1],
        title=title,
        labels={df.columns[0]: x_label, df.columns[1]: y_label},
        color_discrete_sequence=[color]
    )
    fig.update_layout(
        autosize=False,
        width=600,
        height=600
    )
    return fig

# Create forecast charts
def create_forecast_chart(df, title, x_label, y_label, color):
    df[x_label] = pd.to_numeric(df[x_label], errors='coerce')
    df[y_label] = pd.to_numeric(df[y_label], errors='coerce')
    
    # Linear Regression for Forecasting
    X = df[[x_label]].values.reshape(-1, 1)
    y = df[y_label].values
    model = LinearRegression().fit(X, y)
    
    # Forecast for 2030
    future_years = np.array([2030]).reshape(-1, 1)
    forecast_values = model.predict(future_years)
    
    fig = px.line(
        df,
        x=df.columns[0],
        y=df.columns[1],
        title=title,
        labels={df.columns[0]: x_label, df.columns[1]: y_label},
        color_discrete_sequence=[color]
    )
    forecast_df = pd.DataFrame({x_label: [2030], y_label: forecast_values})
    fig.add_scatter(x=forecast_df[x_label], y=forecast_df[y_label], mode='markers+text', text=['2030'], textposition='top center', marker=dict(color='red', size=10))
    fig.update_layout(
        autosize=False,
        width=600,
        height=600
    )
    return fig

# Define custom colors for each chart
colors = ['blue', 'green', 'orange', 'purple']

# Create figures for each sheet
bar_figures = {}
line_figures = {}
forecast_figures = {}

for i, sheet_name in enumerate(sheet_names):
    color = colors[i % len(colors)]
    df = dfs[sheet_name]
    title = titles[sheet_name]
    x_label = x_labels[sheet_name]
    y_label = y_labels[sheet_name]
    
    if sheet_name in ['Sheet1', 'Sheet2', 'Sheet3']:
        bar_figures[sheet_name] = create_bar_chart(df, title, x_label, y_label, color)
        forecast_figures[sheet_name] = create_forecast_chart(df, f"Forecast for 2030 - {title}", x_label, y_label, color)
    elif sheet_name == 'Sheet4':
        line_figures[sheet_name] = create_line_chart(df, title, x_label, y_label, color)

# Load data for Sheet5
df_sheet5 = pd.read_excel(file_path, sheet_name='Sheet5')

# Extract year columns and data for Sheet5
years = df_sheet5.columns[1:10]
departments = df_sheet5.iloc[:, 0]
data = df_sheet5.iloc[:, 1:10]

# Convert data to appropriate format for Sheet5
df_data = data.copy()
df_data.columns = years
df_data['Department'] = departments.values
df_data = df_data.melt(id_vars='Department', var_name='Year', value_name='Percentage')

# Extract 2014 and 2022 data
df_2014 = df_data[df_data['Year'] == '2014'][['Department', 'Percentage']].set_index('Department')
df_2022 = df_data[df_data['Year'] == '2022'][['Department', 'Percentage']].set_index('Department')

# Calculate the difference between 2022 and 2014
percentage_diff = df_2022.join(df_2014, lsuffix='_2022', rsuffix='_2014')
percentage_diff['Percentage Difference'] = percentage_diff['Percentage_2022'] - percentage_diff['Percentage_2014']

# Reset index and prepare data for the graph
percentage_diff = percentage_diff.reset_index()
percentage_diff.columns = ['Department', 'Percentage 2022', 'Percentage 2014', 'Percentage Difference']

# Create the bar graph showing the percentage difference
fig_diff = px.bar(
    percentage_diff,
    x='Department',
    y='Percentage Difference',
    title='Percentage Difference (2022 vs. 2014) for Academic Staff with PhD',
    labels={'Percentage Difference': 'Percentage Difference'},
    height=600,
    color='Department'
)
fig_diff.update_traces(text=percentage_diff['Percentage Difference'], textposition='outside')
fig_diff.update_layout(
    legend_title='Departments',
    autosize=False,
    width=800,
    height=600
)

# Create the bar graph for Sheet5 without text labels
fig_sheet5 = px.bar(
    df_data,
    x='Year',
    y='Percentage',
    color='Department',
    title='Percentage of Full-Time Permanent Academic Staff with PhD (2014-2022)',
    labels={'Percentage': 'Percentage'},
    height=600
)
fig_sheet5.update_traces(textposition='none')  # Remove text annotations

# Define the layout of the app with graphs
app.layout = html.Div(style={'textAlign': 'center'}, children=[
    html.H1("SOFA5 DASHBOARD"),
    
    html.Img(src='/assets/my_image.png', style={'width': '20%', 'height': 'auto'}),
    
    html.Div([
        html.H2("Faculty Data"),
        html.P("Reporting Period 2014 till 2022", style={'color': 'purple'})
    ], style={'textAlign': 'center'}),
    
    # Display the bar charts and forecast charts side by side
    html.Div([
        dcc.Graph(figure=bar_figures.get('Sheet1'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=bar_figures.get('Sheet2'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=bar_figures.get('Sheet3'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=forecast_figures.get('Sheet1'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=forecast_figures.get('Sheet2'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=forecast_figures.get('Sheet3'), style={'width': '50%', 'height': '600px'})
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=line_figures.get('Sheet4'), style={'width': '50%', 'height': '600px'}),
        dcc.Graph(figure=fig_diff, style={'width': '50%', 'height': '600px'})  # Ensure Sheet5 chart is included
    ], style={'display': 'flex', 'justifyContent': 'center'}),
    
    html.Div([
        dcc.Graph(figure=fig_sheet5, style={'width': '100%', 'height': '600px'})  # Full-width for Sheet5
    ], style={'display': 'flex', 'justifyContent': 'center'})
])

# Run the Dash app
if __name__ == '__main__':
    app.run_server(debug=True)
